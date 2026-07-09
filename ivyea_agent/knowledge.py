"""领域知识（方法论摘要）+ 内置知识包检索。

P1.5 起改为从 GBrain (amazon-ops/*) 检索注入；现在先内置精炼版，保证 LLM
复核与用户方法论对齐。
"""
from __future__ import annotations

import json
import re
import hashlib
import sqlite3
import difflib
import html
import io
import zipfile
from importlib import resources
from pathlib import Path
import time
from datetime import datetime
from typing import Any

from . import config, locking, security


class KnowledgeConflictError(RuntimeError):
    """Raised when a reviewed draft no longer matches the current card."""

ALIASES = {
    "否词": ["negative", "negative targeting", "negative keywords"],
    "否定": ["negative", "negative targeting"],
    "预算": ["budget", "daily budget"],
    "出价": ["bid", "bidding"],
    "竞价": ["bid", "bidding"],
    "搜索词": ["search term", "shopping query"],
    "关键词": ["keyword", "keywords"],
    "匹配": ["match", "match types", "broad", "phrase", "exact"],
    "广泛": ["broad"],
    "词组": ["phrase"],
    "精准": ["exact"],
    "listing": ["listing", "product detail page"],
    "详情页": ["listing", "product detail page"],
    "主图": ["images", "product images"],
    "五点": ["bullet", "bullet points"],
    "转化": ["conversion", "cvr"],
    "acos": ["acos", "roas"],
    "赢家词": ["winner", "winning term", "harvest"],
    "收割": ["harvest", "graduate", "manual exact", "manual phrase"],
    "放量": ["scale", "scaling"],
    "扩量": ["scale", "scaling"],
    "承接": ["listing", "conversion", "product detail page"],
    "新品": ["launch", "new product", "automatic targeting", "discovery"],
    "成熟": ["mature", "mature asin", "efficiency"],
    "报表": ["reports", "search term report", "targeting report", "placement report"],
    "位置": ["placement", "top of search", "product pages"],
    "素材": ["content assets", "content_conversion_assets", "content", "images", "video", "a-plus"],
    "a+": ["a-plus", "A+ Content"],
    "高点击": ["clicks", "high clicks"],
    "零单": ["no orders", "0 orders", "no sales"],
    "无订单": ["no orders", "0 orders"],
    "来源": ["source", "source quality", "confidence", "freshness"],
    "置信": ["confidence", "source quality"],
    "时效": ["freshness", "retrieved_at", "version"],
    "注册": ["registration", "seller registration", "account setup", "identity verification"],
    "身份验证": ["identity verification", "verification", "documents", "registration"],
    "验证失败": ["verification error", "registration error", "identity verification"],
    "上架": ["listing", "listings items", "create product listing", "product type definitions"],
    "报错": ["error", "issue", "troubleshooting", "error code"],
    "错误码": ["error code", "issue code", "troubleshooting"],
    "必填属性": ["required attributes", "product type definitions", "listing errors"],
    "绩效": ["account health", "performance", "policy compliance"],
    "账户状况": ["account health", "policy compliance"],
    "停用": ["suspension", "deactivation", "account health", "appeal"],
    "申诉": ["appeal", "plan of action", "account health"],
    "政策": ["policy", "policies", "compliance"],
    "规则": ["policy", "requirements", "guidelines", "compliance"],
    "合规": ["compliance", "policy", "requirements"],
    "知识产权": ["intellectual property", "trademark", "copyright", "patent"],
    "流量": ["traffic", "discoverability", "ranking", "search"],
    "算法": ["algorithm", "ranking", "discoverability", "inference"],
    "自然排名": ["organic rank", "ranking", "discoverability", "inference"],
    "流量池": ["traffic pool", "algorithm", "operator hypothesis", "evidence"],
    "权重": ["weight", "algorithm", "operator hypothesis", "evidence"],
    "归因": ["attribution", "attribution window", "sales scope", "conversion date"],
    "归因窗口": ["attribution window", "attribution model", "sales scope"],
    "展示量": ["impressions", "ctr", "measurement"],
    "点击率": ["ctr", "click-through rate", "impressions", "clicks"],
    "点击成本": ["cpc", "cost per click", "spend", "clicks"],
    "转化率": ["cvr", "conversion rate", "attributed orders", "clicks"],
    "广告报表": ["ads reporting", "search term report", "targeting report", "placement report"],
    "搜索词报告": ["search term report", "click-filter", "inferred search term", "asin"],
    "广告位": ["placement", "top of search", "rest of search", "product pages"],
    "竞价策略": ["bidding strategy", "dynamic bidding", "placement adjustment", "effective bid"],
    "英国站": ["UK", "Europe", "seller registration", "en-GB"],
    "欧洲站": ["EU", "Europe", "seller registration", "VAT"],
    "日本站": ["JP", "Japan", "seller registration", "Amazon.co.jp"],
    "账户健康": ["account health", "account status", "performance notification"],
    "绩效通知": ["performance notification", "account health", "appeal"],
    "账户停用": ["deactivation", "account status changed", "appeal"],
    "受限商品": ["restricted products", "approval", "product safety"],
    "危险品": ["dangerous goods", "hazmat", "SDS", "FBA review"],
    "危品": ["dangerous goods", "hazmat", "SDS"],
    "材料安全数据表": ["SDS", "safety data sheet", "dangerous goods"],
    "费用": ["selling fees", "referral fee", "fee estimate", "product fees"],
    "佣金": ["referral fee", "selling fees"],
    "gtin": ["GTIN", "product ID", "UPC", "EAN", "JAN", "exemption"],
    "条码豁免": ["GTIN exemption", "product ID exemption"],
    "变体": ["variations", "parent-child", "parentage level", "variation theme"],
    "父子体": ["parent-child", "parent SKU", "child", "variations"],
    "知识产权投诉": ["intellectual property complaint", "trademark", "copyright", "patent", "appeal"],
    "税务": ["tax", "VAT", "GST", "sales tax", "tax report"],
    "结算": ["settlement", "payment", "financial transaction", "reconciliation"],
    "对账": ["reconciliation", "settlement", "transaction", "payment"],
    "退货": ["returns", "RMA", "refund", "return report"],
    "退款": ["refund", "returns", "financial event"],
    "索赔": ["claim", "SAFE-T", "reimbursement", "A-to-z"],
    "品牌备案": ["Brand Registry", "trademark", "enrollment", "brand owner"],
    "透明计划": ["Transparency", "serialization", "authenticity", "counterfeit"],
    "展示广告": ["display ads", "Sponsored Display", "audience", "vCPM"],
    "程序化广告": ["Amazon DSP", "programmatic", "supply", "audience"],
    "营销云": ["Amazon Marketing Cloud", "AMC", "clean room", "privacy"],
}

_HIGH_RISK_TERMS = (
    "注册", "身份验证", "验证失败", "上架报错", "报错", "错误码", "error code",
    "绩效", "账户状况", "account health", "停用", "封号", "suspension", "deactivation",
    "申诉", "appeal", "政策", "规则", "合规", "知识产权", "侵权", "费用", "fee",
    "限制", "restricted", "受限商品", "危险品", "危品", "hazmat", "sds", "税务", "vat",
    "gst", "消费税", "jct", "インボイス", "适格请求书", "适格請求書",
    "gtin", "条码豁免", "知识产权投诉", "税务", "结算", "对账", "退款", "索赔", "safe-t",
)
_AMAZON_DOMAIN_TERMS = (
    "amazon", "亚马逊", "seller central", "sp-api", "asin", "fba", "listing", "广告",
    "sponsored products", "sponsored brands", "sponsored display", "acos", "roas", "ctr", "cpc", "cvr",
    "attribution", "placement", "否词", "关键词", "搜索词", "竞价", "出价",
    "预算", "流量", "算法", "排名", "转化", "主图", "五点", "a+", "库存", "店铺",
    "卖家", "上架", "绩效", "注册", "报错", "错误码", "高点击", "零单", "点击", "订单",
    "account health", "product type", "英国站", "欧洲站", "日本站", "受限商品", "危险品",
    "危品", "gtin", "条码豁免", "变体", "父子体", "佣金", "sku", "upc", "ean", "parent_sku",
    "归因", "广告报表", "搜索词报告", "广告位", "竞价策略", "自然排名", "流量池", "权重",
    "税务", "gst", "消费税", "jct", "インボイス", "适格请求书", "适格請求書",
    "结算", "对账", "退货", "退款", "索赔", "safe-t", "brand registry", "品牌备案",
    "透明计划", "transparency", "展示广告", "display ads", "amazon dsp", "程序化广告", "amc", "营销云",
)

METHODOLOGY = """\
你是亚马逊广告运营专家，遵循以下方法论（用户长期沉淀）：

目标底线：目标 ACoS ≤ 毛利率；健康 TACoS 10-15%。
优化优先级链（动作排序铁律）：CTR/CVR 杠杆 > 长尾词辅推 > 出价调整 > 否词。

搜索词分类（每词至少一主标签）：品牌词 / 竞品词 / ASIN串号词 / 核心品类词 / 属性词 / 场景词 / 无关词 / 不确定。
动作标签：放量 / 维持测试 / 降bid / 否词候选 / 观察 / Listing反馈 / 人工复核。

否词规则：以 CPA 为首要标尺（CPA>单品利润进考察，二次分析不直接否）；≥15点击0单或高花费0单→控成本/否候选；不建议词根否定（易误伤）；对比近7天 vs 历史30/60天。
不能否：品牌词、竞品词、战略大词、新品期差词、数据不足词、高转化低流量词、疑似 Listing 承接问题而非流量问题的词。

小类目核心词保护：小类目核心大词即使 ACOS 100%+ 也不降 bid/不否（除非语义过宽的无关宽词）。根因是 CTR 低(主图)+CVR 低(Listing/Review)，靠 主图改版/位置溢价(ToS +30~100%, PP归零)/Down only/拆ToS-only守位/副图bullet评论 解决。

异常归因（低效词不要只说"CVR低"）：无关流量 / Listing承接不足 / 价格或转化问题 / 信号混杂 / 需人工看Listing。

护栏铁律：不投 SBV；不走 Vine；拒绝评论操控；不删 campaign；单次降 bid ≤20%；调整后留 5 天稳定期；不否同义词；不预设产品配置（缺信息写"未指定"）。

证据标签：结论绑证据，区分 [报告]（来自搜索词报表数据）与 [推断]（基于现有信息的判断），不把猜测写成事实。
"""

SOURCE_WATCHLIST = [
    {
        "id": "amazon_ads.sponsored_products",
        "title": "Amazon Ads Sponsored Products",
        "source_type": "official",
        "url": "https://advertising.amazon.com/solutions/products/sponsored-products",
        "category": "amazon_ads",
        "tags": ["sponsored-products", "campaign", "targeting", "budget", "bidding"],
        "license": "amazon_public_docs_summary",
        "priority": 100,
        "review_note": "广告产品、投放入口、展示位置、预算和报告能力的官方基准来源。",
    },
    {
        "id": "amazon_ads.api_docs",
        "title": "Amazon Ads API Documentation",
        "source_type": "official",
        "url": "https://advertising.amazon.com/API/docs/en-us/get-started/overview",
        "category": "amazon_ads_api",
        "tags": ["ads-api", "reporting", "campaign-management", "automation"],
        "license": "amazon_public_docs_summary",
        "priority": 95,
        "review_note": "广告 API 能力、认证、报表和自动化动作的官方入口。",
    },
    {
        "id": "amazon_sp_api.docs",
        "title": "Amazon Selling Partner API Documentation",
        "source_type": "official",
        "url": "https://developer-docs.amazon.com/sp-api/",
        "category": "sp_api",
        "tags": ["sp-api", "catalog", "inventory", "orders", "reports"],
        "license": "amazon_public_docs_summary",
        "priority": 95,
        "review_note": "Seller/Vendor API、模型、Release Notes 和授权流程的官方入口。",
    },
    {
        "id": "amazon_seller.a_plus_content",
        "title": "Amazon A+ Content",
        "source_type": "official",
        "url": "https://sell.amazon.com/tools/a-content",
        "category": "listing",
        "tags": ["a-plus", "listing", "conversion", "content"],
        "license": "amazon_public_docs_summary",
        "priority": 85,
        "review_note": "Listing 承接、A+ 内容和转化素材判断的官方来源。",
    },
    {
        "id": "amazon_seller.fba",
        "title": "Fulfillment by Amazon",
        "source_type": "official",
        "url": "https://sell.amazon.com/fulfillment-by-amazon",
        "category": "inventory",
        "tags": ["fba", "inventory", "fulfillment", "stockout"],
        "license": "amazon_public_docs_summary",
        "priority": 75,
        "review_note": "库存、履约、断货风险与广告放量联动判断的官方来源。",
    },
    {
        "id": "amazon_ads.blog",
        "title": "Amazon Ads Blog / Guides",
        "source_type": "official_plus_blog",
        "url": "https://advertising.amazon.com/library/guides",
        "category": "amazon_ads",
        "tags": ["guide", "launch", "optimization", "case-study"],
        "license": "amazon_public_docs_summary",
        "priority": 70,
        "review_note": "官方指南和案例可辅助方法论，但导入时要与产品帮助页交叉验证。",
    },
    {
        "id": "amazon_seller_forums",
        "title": "Amazon Seller Forums",
        "source_type": "community_official_forum",
        "url": "https://sellercentral.amazon.com/seller-forums",
        "category": "community",
        "tags": ["seller-forum", "policy", "operations", "case"],
        "license": "community_summary_requires_review",
        "priority": 55,
        "review_note": "适合发现真实运营问题和边界案例；不能覆盖官方规则，必须人工复核。",
    },
    {
        "id": "zhiwubuyan",
        "title": "知无不言跨境电商社区",
        "source_type": "community",
        "url": "https://www.wearesellers.com/",
        "category": "community",
        "tags": ["community", "seller-case", "china-seller", "operations"],
        "license": "community_summary_requires_review",
        "priority": 45,
        "review_note": "适合沉淀中文卖家经验和案例；导入前必须去广告化、去个人隐私、去未经验证结论。",
    },
]

IMPORTABLE_SUFFIXES = {
    ".md", ".markdown", ".txt", ".csv", ".tsv", ".json", ".yaml", ".yml",
    ".html", ".htm", ".docx", ".xlsx", ".xlsm", ".pdf",
}
IGNORED_IMPORT_DIRS = {".git", ".hg", ".svn", "__pycache__", "node_modules", ".venv", "venv"}
DEFAULT_IMPORT_FILE_BYTES = 5 * 1024 * 1024


def _base():
    return resources.files("ivyea_agent").joinpath("knowledge_base")


def _user_base() -> Path:
    return config.IVYEA_DIR / "knowledge"


def _sources_file() -> Path:
    return _user_base() / "sources.jsonl"


def _index_file() -> Path:
    return _user_base() / "index.db"


def _uploads_dir() -> Path:
    return _user_base() / "uploads"


def _upload_history_file() -> Path:
    return _user_base() / "uploads.jsonl"


def _mutation_lock_file() -> Path:
    return _user_base() / ".mutation.lock"


def _versions_file() -> Path:
    return _user_base() / "versions.jsonl"


def _versions_dir() -> Path:
    return _user_base() / "versions"


def list_cards() -> list[dict[str, Any]]:
    """Return bundled and user knowledge card metadata."""
    text = _base().joinpath("index.json").read_text(encoding="utf-8")
    cards = json.loads(text)
    for card in cards:
        _enrich_builtin_card(card)
    cards.extend(list_user_cards())
    return cards


def list_builtin_cards() -> list[dict[str, Any]]:
    text = _base().joinpath("index.json").read_text(encoding="utf-8")
    rows = json.loads(text)
    for card in rows:
        _enrich_builtin_card(card)
    return rows


def _enrich_builtin_card(card: dict[str, Any]) -> dict[str, Any]:
    card.setdefault("retrieved_at", card.get("version", ""))
    card.setdefault("confidence", _confidence(card.get("source_type", "")))
    card.setdefault("freshness", _freshness(card))
    card.setdefault("source_quality", _source_quality(card))
    card.setdefault("license", "amazon_public_docs_summary")
    card.setdefault("scope", "builtin")
    card.setdefault("authority_tier", _authority_tier(card))
    card.setdefault("evidence_class", _evidence_class(card))
    card.setdefault("marketplaces", ["GLOBAL"])
    card.setdefault("locales", ["en-US", "zh-CN"])
    try:
        body = _base().joinpath(card["path"]).read_text(encoding="utf-8")
        card.setdefault("body_hash", _hash(body))
    except (KeyError, OSError, UnicodeDecodeError):
        card.setdefault("body_hash", "")
    return card


def list_user_cards() -> list[dict[str, Any]]:
    p = _sources_file()
    if not p.exists():
        return []
    rows = []
    for line in p.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            card = json.loads(line)
        except Exception:
            continue
        card.setdefault("source_type", "user")
        card.setdefault("confidence", _confidence(card.get("source_type", "")))
        card.setdefault("retrieved_at", "")
        card.setdefault("freshness", _freshness(card))
        card.setdefault("source_quality", _source_quality(card))
        card.setdefault("license", "user_supplied")
        if not card.get("body_hash"):
            try:
                card["body_hash"] = _hash(_user_base().joinpath(card["path"]).read_text(encoding="utf-8"))
            except Exception:
                card["body_hash"] = ""
        card.setdefault("scope", "user")
        card.setdefault("authority_tier", _authority_tier(card))
        card.setdefault("evidence_class", _evidence_class(card))
        card.setdefault("marketplaces", ["ACCOUNT_LOCAL"])
        card.setdefault("locales", ["USER_SUPPLIED"])
        rows.append(card)
    return rows


def _confidence(source_type: str) -> str:
    if source_type == "official":
        return "high"
    if source_type.startswith("official_plus"):
        return "medium_high"
    if source_type.startswith("community"):
        return "medium"
    if source_type == "user":
        return "user_supplied"
    if source_type == "account_authorized_official_evidence":
        return "account_observed"
    if source_type.startswith("internal"):
        return "high_control_only"
    return "unknown"


def _source_quality(card: dict[str, Any]) -> str:
    source_type = str(card.get("source_type") or "")
    scope = str(card.get("scope") or "")
    if source_type == "official":
        return "authoritative"
    if source_type.startswith("official_plus"):
        return "synthesized_with_official_anchor"
    if source_type.startswith("community"):
        return "directional_requires_account_validation"
    if source_type == "account_authorized_official_evidence":
        return "account_observed_official_context"
    if source_type == "legacy_gbrain":
        return "operator_local_requires_official_or_account_validation"
    if scope == "user" or source_type == "user":
        return "account_local_overrides_generic_knowledge"
    if source_type.startswith("internal"):
        return "internal_control_not_external_evidence"
    return "unknown_requires_review"


def _authority_tier(card: dict[str, Any]) -> str:
    source_type = str(card.get("source_type") or "")
    scope = str(card.get("scope") or "")
    if source_type == "official":
        return "primary"
    if source_type.startswith("official_plus"):
        return "secondary_synthesis"
    if source_type.startswith("community"):
        return "community_directional"
    if source_type == "account_authorized_official_evidence":
        return "account_local"
    if source_type == "legacy_gbrain":
        return "operator_local"
    if scope == "user" or source_type == "user":
        return "account_local"
    if source_type.startswith("internal"):
        return "internal_governance"
    return "unclassified"


def _evidence_class(card: dict[str, Any]) -> str:
    source_type = str(card.get("source_type") or "")
    category = str(card.get("category") or "")
    if source_type == "official":
        if category == "policies":
            return "official_policy_summary"
        return "official_documentation_summary"
    if source_type.startswith("official_plus"):
        return "official_anchored_operating_synthesis"
    if source_type.startswith("community"):
        return "operator_hypothesis"
    if source_type == "account_authorized_official_evidence":
        return "account_authorized_official_evidence"
    if source_type == "legacy_gbrain":
        return "operator_hypothesis"
    if card.get("scope") == "user" or source_type == "user":
        return "account_local_evidence"
    return "unclassified"


def _authority_score(card: dict[str, Any]) -> int:
    tier = str(card.get("authority_tier") or _authority_tier(card))
    return {
        "primary": 12,
        "primary_dynamic": 12,
        "account_local": 10,
        "secondary_synthesis": 6,
        "internal_governance": 4,
        "community_directional": 1,
        "operator_local": 2,
    }.get(tier, 0)


def _freshness(card: dict[str, Any]) -> str:
    stamp = str(card.get("retrieved_at") or card.get("version") or "").strip()
    if not stamp:
        return "undated"
    parsed = None
    for fmt in ("%Y-%m-%d", "%Y.%m", "%Y-%m", "%Y"):
        try:
            parsed = datetime.strptime(stamp, fmt)
            break
        except ValueError:
            continue
    if parsed is None:
        return "reviewed"
    now = datetime.fromtimestamp(time.time())
    months = (now.year - parsed.year) * 12 + (now.month - parsed.month)
    if months <= 6:
        return "current"
    if months <= 12:
        return "aging_review_soon"
    return "stale_needs_review"


def get_card(card_id: str) -> dict[str, Any] | None:
    for card in list_cards():
        if card["id"] == card_id:
            body = _read_body(card)
            return {**card, "body": body}
    return None


def _read_body(card: dict[str, Any]) -> str:
    if card.get("scope") == "user":
        p = _user_base().joinpath(card["path"])
        return p.read_text(encoding="utf-8")
    return _base().joinpath(card["path"]).read_text(encoding="utf-8")


def _score(text: str, terms: list[str]) -> int:
    low = text.lower()
    return sum(low.count(t.lower()) for t in terms if t)


def _hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="replace")).hexdigest()


def search(query: str, limit: int = 5) -> list[dict[str, Any]]:
    """Simple deterministic search over bundled and user knowledge cards."""
    terms = re.findall(r"[\w\u4e00-\u9fff+.-]+", query)
    expanded = []
    for t in terms:
        expanded.append(t)
        expanded.extend(ALIASES.get(t.lower(), []))
        expanded.extend(ALIASES.get(t, []))
    for key, vals in ALIASES.items():
        if key in query:
            expanded.append(key)
            expanded.extend(vals)
    terms = expanded
    rows = []
    for card in list_cards():
        body = _read_body(card)
        hay = " ".join([card["id"], card["title"], " ".join(card.get("tags", [])), body])
        lexical_score = _score(hay, terms)
        if lexical_score:
            score = lexical_score * 10 + _authority_score(card)
            snippet = _snippet(body, terms)
            rows.append({
                **card,
                "score": score,
                "lexical_score": lexical_score,
                "authority_score": _authority_score(card),
                "snippet": snippet,
            })
    rows.sort(key=lambda r: (-r["score"], r["id"]))
    return rows[:limit]


def _snippet(body: str, terms: list[str], width: int = 220) -> str:
    low = body.lower()
    pos = -1
    for t in terms:
        pos = low.find(t.lower())
        if pos >= 0:
            break
    if pos < 0:
        return body[:width].replace("\n", " ").strip()
    start = max(0, pos - width // 3)
    end = min(len(body), start + width)
    return body[start:end].replace("\n", " ").strip()


def render_search(query: str, limit: int = 5) -> str:
    hits = search(query, limit=limit)
    if not hits:
        return "（无匹配知识）"
    lines = []
    for idx, h in enumerate(hits, 1):
        source = f" · {h['source_url']}" if h.get("source_url") else ""
        meta = (
            f"{h['source_type']} confidence={h.get('confidence', 'unknown')} "
            f"freshness={h.get('freshness', '-')} quality={h.get('source_quality', '-')}"
        )
        lines.append(f"- [K{idx}] {h['id']} · {h['title']} [{meta}]{source}\n  {h['snippet']}")
    return "\n".join(lines)


def render_audit() -> str:
    """Show source quality metadata for bundled and user knowledge cards."""
    lines = ["Ivyea 知识库审计："]
    for card in audit()["cards"]:
        source = card.get("source_url") or "-"
        lines.append(
            f"- {card['id']} | {card.get('scope', 'builtin')} | {card['source_type']} | confidence={card.get('confidence')} | "
            f"freshness={card.get('freshness')} | quality={card.get('source_quality')} | "
            f"retrieved={card.get('retrieved_at')} | license={card.get('license', '-')} | hash={str(card.get('body_hash', ''))[:12]} | source={source}"
        )
    return "\n".join(lines)


def source_registry() -> dict[str, Any]:
    """Return a product-facing source registry grouped by URL/type/license."""
    cards = list_cards()
    grouped: dict[str, dict[str, Any]] = {}
    for card in cards:
        source_url = str(card.get("source_url") or "").strip()
        source_type = str(card.get("source_type") or "unknown")
        license_name = str(card.get("license") or "unknown")
        category = str(card.get("category") or "uncategorized")
        scope = str(card.get("scope") or "builtin")
        key = source_url or f"{scope}:{source_type}:{category}:{license_name}"
        row = grouped.setdefault(key, {
            "key": key,
            "source_url": source_url,
            "source_type": source_type,
            "scope": scope,
            "license": license_name,
            "categories": set(),
            "cards": [],
            "card_count": 0,
            "stale_cards": 0,
            "missing_source_url": not bool(source_url),
            "confidence_levels": set(),
            "freshness_levels": set(),
            "source_qualities": set(),
        })
        row["categories"].add(category)
        row["confidence_levels"].add(str(card.get("confidence") or "unknown"))
        row["freshness_levels"].add(str(card.get("freshness") or "unknown"))
        row["source_qualities"].add(str(card.get("source_quality") or "unknown"))
        row["cards"].append({
            "id": card.get("id", ""),
            "title": card.get("title", ""),
            "category": category,
            "freshness": card.get("freshness", ""),
            "confidence": card.get("confidence", ""),
            "body_hash": card.get("body_hash", ""),
            "authority_tier": card.get("authority_tier", ""),
            "evidence_class": card.get("evidence_class", ""),
            "marketplaces": list(card.get("marketplaces") or []),
            "locales": list(card.get("locales") or []),
            "evidence_id": card.get("evidence_id", ""),
        })
        row["card_count"] += 1
        if card.get("freshness") == "stale_needs_review":
            row["stale_cards"] += 1

    sources = []
    for row in grouped.values():
        item = dict(row)
        item["categories"] = sorted(row["categories"])
        item["confidence_levels"] = sorted(row["confidence_levels"])
        item["freshness_levels"] = sorted(row["freshness_levels"])
        item["source_qualities"] = sorted(row["source_qualities"])
        item["review_required"] = bool(item["stale_cards"]) or (
            bool(item["missing_source_url"]) and item["source_type"].startswith("community")
        )
        sources.append(item)
    sources.sort(key=lambda r: (r["scope"] != "builtin", r["source_type"], r["key"]))
    summary = {
        "cards": len(cards),
        "sources": len(sources),
        "official_sources": len([s for s in sources if s["source_type"] == "official"]),
        "community_sources": len([s for s in sources if str(s["source_type"]).startswith("community")]),
        "user_sources": len([s for s in sources if s["scope"] == "user" or s["source_type"] == "user"]),
        "missing_source_url_cards": len([c for c in cards if not c.get("source_url")]),
        "stale_sources": len([s for s in sources if s["stale_cards"]]),
        "review_required_sources": len([s for s in sources if s["review_required"]]),
        "licenses": _counts([str(c.get("license") or "unknown") for c in cards]),
        "categories": _counts([str(c.get("category") or "uncategorized") for c in cards]),
    }
    return {"summary": summary, "sources": sources}


def render_source_registry() -> str:
    data = source_registry()
    s = data["summary"]
    lines = [
        "Ivyea 知识来源登记表：",
        f"- cards={s['cards']} sources={s['sources']} official={s['official_sources']} "
        f"community={s['community_sources']} user={s['user_sources']} review_required={s['review_required_sources']}",
    ]
    for source in data["sources"]:
        url = source.get("source_url") or "(no source_url)"
        flags = []
        if source.get("missing_source_url"):
            flags.append("missing-url")
        if source.get("review_required"):
            flags.append("review")
        flag_text = f" [{' '.join(flags)}]" if flags else ""
        categories = ",".join(source.get("categories") or [])
        card_ids = ",".join(c.get("id", "") for c in source.get("cards") or [])
        lines.append(
            f"- {source['source_type']} | {source['scope']} | {source['license']} | "
            f"cards={source['card_count']} | categories={categories}{flag_text}\n"
            f"  {url}\n"
            f"  ids={card_ids}"
        )
    return "\n".join(lines)


def source_watchlist() -> dict[str, Any]:
    """Return curated sources that IvyeaAgent should monitor/import from with review."""
    rows = []
    for source in SOURCE_WATCHLIST:
        row = dict(source)
        source_type = str(row.get("source_type") or "")
        row["review_required"] = (
            source_type.startswith("community")
            or row.get("license") == "community_summary_requires_review"
        )
        row["confidence"] = _confidence(source_type)
        rows.append(row)
    rows.sort(key=lambda r: (-int(r.get("priority") or 0), r.get("id", "")))
    summary = {
        "sources": len(rows),
        "official_sources": len([r for r in rows if r["source_type"] == "official"]),
        "community_sources": len([r for r in rows if str(r["source_type"]).startswith("community")]),
        "review_required_sources": len([r for r in rows if r.get("review_required")]),
        "categories": _counts([str(r.get("category") or "uncategorized") for r in rows]),
        "policy": "manual_review_before_import",
    }
    return {"summary": summary, "sources": rows}


def render_source_watchlist() -> str:
    data = source_watchlist()
    s = data["summary"]
    lines = [
        "Ivyea Amazon 知识来源观察清单：",
        f"- sources={s['sources']} official={s['official_sources']} "
        f"community={s['community_sources']} review_required={s['review_required_sources']} "
        f"policy={s['policy']}",
    ]
    for source in data["sources"]:
        review = " review" if source.get("review_required") else ""
        tags = ",".join(source.get("tags") or [])
        lines.append(
            f"- {source['id']} | {source['source_type']} | priority={source.get('priority', 0)}{review}\n"
            f"  {source['title']}\n"
            f"  {source['url']}\n"
            f"  tags={tags}\n"
            f"  note={source.get('review_note', '')}"
        )
    return "\n".join(lines)


def _counts(values: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    for value in values:
        out[value] = out.get(value, 0) + 1
    return dict(sorted(out.items(), key=lambda item: item[0]))


def audit() -> dict[str, Any]:
    """Return structured source quality metadata for product integrations."""
    cards = []
    for card in list_cards():
        cards.append({
            "id": card.get("id", ""),
            "title": card.get("title", ""),
            "category": card.get("category", ""),
            "scope": card.get("scope", "builtin"),
            "source_type": card.get("source_type", ""),
            "confidence": card.get("confidence", ""),
            "freshness": card.get("freshness", ""),
            "source_quality": card.get("source_quality", ""),
            "retrieved_at": card.get("retrieved_at", ""),
            "license": card.get("license", ""),
            "source_url": card.get("source_url", ""),
            "tags": list(card.get("tags") or []),
            "body_hash": card.get("body_hash", ""),
            "authority_tier": card.get("authority_tier", ""),
            "evidence_class": card.get("evidence_class", ""),
            "marketplaces": list(card.get("marketplaces") or []),
            "locales": list(card.get("locales") or []),
            "evidence_id": card.get("evidence_id", ""),
            "evidence_kind": card.get("evidence_kind", ""),
            "observed_at": card.get("observed_at", ""),
            "diagnostic": card.get("diagnostic") or {},
        })
    conflict_rows = conflicts()
    registry = source_registry()
    summary = {
        "cards": len(cards),
        "builtin_cards": len([c for c in cards if c.get("scope") != "user"]),
        "user_cards": len([c for c in cards if c.get("scope") == "user"]),
        "official_cards": len([c for c in cards if str(c.get("source_type") or "") == "official"]),
        "stale_cards": len([c for c in cards if c.get("freshness") == "stale_needs_review"]),
        "conflicts": len(conflict_rows),
        "source_registry": registry["summary"],
        "sources": str(_sources_file()),
        "index": str(_index_file()),
    }
    return {"summary": summary, "cards": cards, "conflicts": conflict_rows, "source_registry": registry}


def context_for_query(query: str, limit: int = 3, max_chars: int = 1200) -> tuple[str, list[str]]:
    """Return compact context snippets for prompt injection plus selected card ids."""
    evidence = evidence_context(query, limit=limit, max_chars=max_chars)
    text = str(evidence.get("text") or "")
    if len(text) > max_chars:
        text = text[:max_chars].rstrip() + "\n..."
    return text, list(evidence.get("ids") or [])


def retrieval_decision(query: str) -> dict[str, Any]:
    """Decide whether Amazon evidence should be injected and how strict to be."""
    low = str(query or "").lower().strip()
    high_matches = sorted({term for term in _HIGH_RISK_TERMS if term in low})
    domain_matches = sorted({term for term in _AMAZON_DOMAIN_TERMS if term in low})
    diagnostic_issue = bool(domain_matches) and any(
        term in low for term in (
            "错误", "失败", "异常", "被拒", "不通过", "报错", "issue", "suppressed", "invalid", "required",
        )
    )
    diagnostic_issue = diagnostic_issue or (
        bool(domain_matches) and bool(re.search(r"\b(?:[a-z]{1,8}\d{3,}|\d{4,})\b", low))
    )
    if diagnostic_issue and not high_matches:
        high_matches = ["amazon_diagnostic_issue"]
    if high_matches:
        return {
            "should_retrieve": True,
            "risk": "high",
            "reason": "policy_or_account_impact",
            "matched_terms": high_matches,
        }
    if domain_matches:
        risk = "medium" if any(term in low for term in (
            "广告", "流量", "算法", "listing", "fba", "转化", "sponsored products", "sponsored brands",
            "sponsored display", "acos", "roas", "ctr", "cpc", "cvr", "归因", "attribution", "搜索词",
            "search term", "广告位", "placement", "campaign manager",
            "展示广告", "display ads", "amazon dsp", "程序化广告", "amc", "营销云", "clean room",
        )) else "low"
        return {
            "should_retrieve": True,
            "risk": risk,
            "reason": "amazon_domain_question",
            "matched_terms": domain_matches,
        }
    return {
        "should_retrieve": False,
        "risk": "none",
        "reason": "no_amazon_domain_signal",
        "matched_terms": [],
    }


def evidence_context(query: str, limit: int = 4, max_chars: int = 2600) -> dict[str, Any]:
    """Return ranked evidence, stable citation keys and a prompt-ready context block."""
    decision = retrieval_decision(query)
    if not decision["should_retrieve"]:
        return {
            **decision, "text": "", "citations": [], "ids": [], "hits": [],
            "freshness_review_required": False,
        }
    requested = max(1, min(int(limit or 4), 10))
    hits = search(query, limit=min(50, requested * 3))
    algorithm_question = any(term in str(query or "").lower() for term in ("算法", "algorithm", "流量池", "权重", "自然排名", "organic rank"))
    query_low = str(query or "").lower()
    specific_terms = [
        term.lower() for term in re.findall(r"[A-Za-z0-9][A-Za-z0-9._-]{5,}", str(query or ""))
        if (re.search(r"[A-Za-z]", term) and (re.search(r"\d", term) or "-" in term or "_" in term))
    ]

    def topic_bonus(hit: dict[str, Any]) -> int:
        category = str(hit.get("category") or "")
        hay = " ".join([
            str(hit.get("id") or ""), str(hit.get("title") or ""), category,
            " ".join(hit.get("tags") or []), str(hit.get("snippet") or ""),
        ]).lower()
        bonus = 0
        for exact in re.findall(r"\b(?:[A-Za-z]{1,8}\d{3,}|\d{4,})\b", str(query or "")):
            if exact.lower() in hay:
                bonus += 200
        category_rules = [
            (("注册", "registration", "身份验证", "verification"), {"seller_registration", "registration_errors"}),
            (("费用", "fee", "佣金", "referral"), {"seller_fees"}),
            (("受限", "restricted"), {"restricted_products"}),
            (("危险品", "危品", "hazmat", "sds"), {"dangerous_goods"}),
            (("gtin", "upc", "ean", "条码"), {"listing_requirements"}),
            (("变体", "variation", "parent_sku", "父子体"), {"listing_requirements"}),
            (("绩效", "account health", "停用", "申诉", "appeal"), {"account_health", "policies"}),
            (("报错", "错误", "error", "invalid", "required"), {"listing_errors", "registration_errors"}),
            (("广告", "acos", "roas", "ctr", "cpc", "cvr", "归因", "attribution"), {"ads_measurement"}),
            (("报表", "报告", "report", "搜索词", "search term", "广告位", "placement"), {"ads_reporting"}),
            (("实验", "测试", "experiment", "因果", "causal"), {"ads_experimentation"}),
            (("算法", "algorithm", "流量池", "权重", "自然排名", "organic rank"), {"traffic_governance"}),
            (("税务", "tax", "vat", "gst", "sales tax", "消费税", "jct", "インボイス", "适格请求书"), {"tax_compliance"}),
            (("结算", "对账", "settlement", "reconciliation", "payment"), {"finance_settlement"}),
            (("退货", "退款", "索赔", "returns", "refund", "safe-t", "claim"), {"returns_claims"}),
            (("品牌备案", "brand registry", "商标", "trademark"), {"brand_registry"}),
            (("透明计划", "transparency", "防伪", "serialization"), {"brand_protection"}),
            (("sponsored brands", "品牌广告"), {"sponsored_brands"}),
            (("展示广告", "display ads", "sponsored display"), {"display_ads"}),
            (("amazon dsp", "程序化广告"), {"amazon_dsp"}),
            (("amazon marketing cloud", "amc", "营销云", "clean room"), {"ads_clean_room"}),
        ]
        for terms, categories in category_rules:
            if category in categories and any(term in query_low for term in terms):
                bonus += 60
        hit_id = str(hit.get("id") or "")
        hit_markets = {str(value).upper() for value in hit.get("marketplaces") or []}
        if any(term in query_low for term in ("日本", "japan", "jp")) and "JP" in hit_markets:
            bonus += 90
        if any(term in query_low for term in ("英国", "uk", "欧洲", "europe", "eu")) and hit_markets.intersection({"UK", "DE", "FR", "IT", "ES", "NL", "SE", "PL"}):
            bonus += 90
        if hit_id == "amazon_ads.bid_stack_and_auction" and any(
            term in query_low for term in ("动态竞价", "dynamic bidding", "叠加", "effective bid")
        ):
            bonus += 100
        return bonus

    def evidence_priority(hit: dict[str, Any]) -> tuple[int, int, int, str]:
        tier = str(hit.get("authority_tier") or "")
        hit_hay = " ".join([
            str(hit.get("id") or ""), str(hit.get("title") or ""), str(hit.get("snippet") or ""),
            " ".join(hit.get("tags") or []),
        ]).lower()
        account_specific = tier == "account_local" and any(term in hit_hay for term in specific_terms)
        if account_specific:
            priority = 130
        elif algorithm_question and hit.get("id") == "governance.traffic_algorithm_evidence":
            priority = 120
        elif algorithm_question and hit.get("id") == "governance.professional_knowledge_standard":
            priority = 115
        elif tier.startswith("primary"):
            priority = 110
        elif tier == "internal_governance":
            priority = 100
        elif tier == "account_local":
            priority = 90
        elif tier == "secondary_synthesis":
            priority = 80
        elif tier in {"operator_local", "community_directional"}:
            priority = 30
        else:
            priority = 10
        freshness = str(hit.get("freshness") or "")
        if freshness == "stale_needs_review":
            priority -= 25
        elif freshness in {"aging_review_soon", "undated"}:
            priority -= 10
        return priority, topic_bonus(hit), int(hit.get("score") or 0), str(hit.get("id") or "")

    hits.sort(key=lambda hit: (
        -evidence_priority(hit)[0], -evidence_priority(hit)[1],
        -evidence_priority(hit)[2], evidence_priority(hit)[3],
    ))
    hits = hits[:requested]
    freshness_review_required = any(
        hit.get("freshness") in {"stale_needs_review", "aging_review_soon", "undated"} for hit in hits
    )
    citations: list[dict[str, Any]] = []
    lines: list[str] = []
    for idx, hit in enumerate(hits, 1):
        key = f"K{idx}"
        citation = {
            "key": key,
            "id": hit["id"],
            "title": hit["title"],
            "url": str(hit.get("source_url") or ""),
            "source_type": hit.get("source_type", "unknown"),
            "authority_tier": hit.get("authority_tier", "unclassified"),
            "evidence_class": hit.get("evidence_class", "unclassified"),
            "freshness": hit.get("freshness", "unknown"),
            "retrieved_at": hit.get("retrieved_at", ""),
            "marketplaces": list(hit.get("marketplaces") or ["GLOBAL"]),
            "locales": list(hit.get("locales") or []),
            "snippet": hit["snippet"],
        }
        citations.append(citation)
        lines.append(
            f"[{key}] {citation['title']} | id={citation['id']} | authority={citation['authority_tier']} | "
            f"evidence={citation['evidence_class']} | confidence={hit.get('confidence', 'unknown')} | "
            f"freshness={citation['freshness']} | marketplace={','.join(citation['marketplaces'])} | "
            f"url={citation['url'] or '(internal/no-url)'}\n"
            f"excerpt: {citation['snippet']}"
        )
    if lines:
        text = (
            f"检索决策：risk={decision['risk']} reason={decision['reason']}。\n"
            + "\n".join(lines)
            + "\n引用规则：仅在结论确实由摘录支持时使用对应 [K#]；官方事实、数据推断、运营假设必须明确区分。"
        )
        if freshness_review_required:
            text += "\n时效门禁：命中证据包含过期、临期或无日期内容；高风险结论必须先核对当前官方来源。"
    else:
        text = (
            f"检索决策：risk={decision['risk']} reason={decision['reason']}，但内部知识库没有命中。"
            "不得把猜测写成亚马逊官方规则；应明确知识缺口，并建议核对对应站点的最新官方页面。"
        )
    if len(text) > max_chars:
        text = text[:max_chars].rstrip() + "\n..."
        visible_keys = set(citation_keys(text))
        visible_pairs = [
            (citation, hit) for citation, hit in zip(citations, hits)
            if citation["key"] in visible_keys
        ]
        citations = [pair[0] for pair in visible_pairs]
        hits = [pair[1] for pair in visible_pairs]
    return {
        **decision,
        "text": text,
        "citations": citations,
        "ids": [hit["id"] for hit in hits],
        "hits": hits,
        "freshness_review_required": freshness_review_required,
    }


def citation_keys(text: str) -> list[str]:
    """Extract unique citation keys in their first-use order."""
    seen: set[str] = set()
    keys: list[str] = []
    for match in re.finditer(r"\[(K\d+)\]", str(text or ""), flags=re.IGNORECASE):
        key = match.group(1).upper()
        if key not in seen:
            seen.add(key)
            keys.append(key)
    return keys


def validate_citations(text: str, citations: list[dict[str, Any]]) -> dict[str, Any]:
    available = {str(row.get("key") or "").upper() for row in citations if row.get("key")}
    used = citation_keys(text)
    valid = [key for key in used if key in available]
    invalid = [key for key in used if key not in available]
    return {
        "ok": bool(valid) and not invalid,
        "available": sorted(available, key=lambda key: int(key[1:]) if key[1:].isdigit() else 0),
        "used": used,
        "valid": valid,
        "invalid": invalid,
        "missing": not bool(valid),
    }


def merge_citations(
    existing: list[dict[str, Any]], incoming: list[dict[str, Any]], text: str,
) -> tuple[list[dict[str, Any]], str]:
    """Merge evidence from repeated searches and remap incoming keys without collisions."""
    merged = [dict(row) for row in existing]
    by_id = {str(row.get("id") or ""): str(row.get("key") or "") for row in merged if row.get("id")}
    numbers = [int(key[1:]) for key in (str(row.get("key") or "") for row in merged)
               if re.fullmatch(r"K\d+", key)]
    next_number = max(numbers, default=0) + 1
    mapping: dict[str, str] = {}
    for raw in incoming:
        row = dict(raw)
        old_key = str(row.get("key") or "").upper()
        card_id = str(row.get("id") or "")
        new_key = by_id.get(card_id, "")
        if not new_key:
            new_key = f"K{next_number}"
            next_number += 1
            row["key"] = new_key
            merged.append(row)
            if card_id:
                by_id[card_id] = new_key
        mapping[old_key] = new_key

    def replace(match: re.Match[str]) -> str:
        old_key = match.group(1).upper()
        return f"[{mapping.get(old_key, old_key)}]"

    rewritten = re.sub(r"\[(K\d+)\]", replace, str(text or ""), flags=re.IGNORECASE)
    return merged, rewritten


def append_citation_footer(text: str, citations: list[dict[str, Any]]) -> str:
    """Append only the sources actually cited by the answer."""
    check = validate_citations(text, citations)
    if not check["valid"] or "\n引用知识：\n" in text:
        return text
    by_key = {str(row.get("key") or "").upper(): row for row in citations}
    lines = ["引用知识："]
    for key in check["valid"]:
        row = by_key[key]
        source = row.get("url") or f"ivyea://knowledge/{row.get('id', '')}"
        lines.append(
            f"- [{key}] {row.get('title') or row.get('id')} — {source} "
            f"({row.get('authority_tier', 'unclassified')}, {row.get('freshness', 'unknown')})"
        )
    return str(text or "").rstrip() + "\n\n" + "\n".join(lines)


def _slug(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip().lower()).strip("-")
    return slug[:80] or time.strftime("knowledge-%Y%m%d-%H%M%S")


def _safe_filename(filename: str) -> str:
    name = Path(str(filename or "upload")).name
    stem = _slug(Path(name).stem or "upload")
    suffix = re.sub(r"[^A-Za-z0-9.]+", "", Path(name).suffix.lower())[:16]
    return f"{stem}{suffix}" if suffix else stem


def _safe_path_segment(segment: str, fallback: str = "item") -> str:
    raw = str(segment or "").strip()
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", raw).strip("-._").lower()
    if cleaned:
        return cleaned[:80]
    return f"{fallback}-{_hash(raw or fallback)[:12]}"


def _first_heading_or_stem(body: str, fallback: str) -> str:
    for line in str(body or "").splitlines()[:20]:
        text = line.strip()
        if text.startswith("#"):
            title = text.lstrip("#").strip()
            if title:
                return title[:160]
    return str(fallback or "导入知识").strip()[:160] or "导入知识"


def _safe_rel_path(path: str) -> Path:
    raw = str(path or "").strip().replace("\\", "/")
    if not raw or raw.startswith("/") or raw.startswith("../") or "/../" in raw or raw == "..":
        raise ValueError("invalid knowledge file path")
    rel = Path(raw)
    if rel.parts and rel.parts[0] not in {"user", "uploads"}:
        raise ValueError("knowledge file path must be under user/ or uploads/")
    return rel


def _resolve_user_path(path: str) -> Path:
    rel = _safe_rel_path(path)
    base = _user_base().resolve()
    target = (base / rel).resolve()
    if target != base and base not in target.parents:
        raise ValueError("knowledge file path escapes knowledge directory")
    return target


def _tag_list(tags: list[str] | str | None) -> list[str]:
    if isinstance(tags, str):
        return [t.strip() for t in tags.split(",") if t.strip()]
    if isinstance(tags, list):
        return [str(t).strip() for t in tags if str(t).strip()]
    return []


def _write_text_atomic(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + f".{time.time_ns()}.tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)


def _write_json_atomic(path: Path, value: Any) -> None:
    _write_text_atomic(path, json.dumps(value, ensure_ascii=False, indent=2) + "\n")


def _version_rows() -> list[dict[str, Any]]:
    path = _versions_file()
    if not path.exists():
        return []
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(row, dict):
            rows.append(row)
    return rows


def _new_version_id(card_id: str, revision: int, body_hash: str) -> str:
    material = f"{card_id}|{revision}|{body_hash}|{time.time_ns()}"
    return "kv-" + hashlib.sha256(material.encode("utf-8")).hexdigest()[:20]


def _record_version_unlocked(
    card: dict[str, Any],
    body: str,
    *,
    action: str,
    actor: str,
    actor_source: str,
    rollback_from: str = "",
) -> dict[str, Any]:
    version_id = str(card.get("current_version_id") or "")
    if not version_id:
        raise ValueError("versioned card is missing current_version_id")
    card_id = str(card.get("id") or "")
    created_at = datetime.now().astimezone().isoformat(timespec="seconds")
    snapshot_path = _versions_dir() / _slug(card_id) / f"{version_id}.json"
    snapshot = {
        "version_id": version_id,
        "card_id": card_id,
        "revision": int(card.get("revision") or 1),
        "created_at": created_at,
        "action": action,
        "actor": security.redact_text(str(actor or "local-operator"))[:120],
        "actor_source": str(actor_source or "local_operation")[:80],
        "rollback_from": str(rollback_from or ""),
        "body_hash": str(card.get("body_hash") or _hash(body)),
        "parent_version_id": str(card.get("parent_version_id") or ""),
        "card": dict(card),
        "body": body,
    }
    _write_json_atomic(snapshot_path, snapshot)
    ledger = {key: value for key, value in snapshot.items() if key not in {"card", "body"}}
    ledger["snapshot"] = str(snapshot_path)
    path = _versions_file()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(ledger, ensure_ascii=False) + "\n")
        fh.flush()
    return ledger


def list_versions(card_id: str = "", limit: int = 100) -> dict[str, Any]:
    """List immutable user-card revisions without exposing snapshot bodies."""
    clean_id = str(card_id or "").strip()
    rows = _version_rows()
    if clean_id:
        rows = [row for row in rows if row.get("card_id") == clean_id]
    rows = rows[-max(1, min(int(limit or 100), 1000)):]
    rows.reverse()
    return {
        "summary": {"versions": len(rows), "card_id": clean_id},
        "versions": rows,
        "ledger": str(_versions_file()),
    }


def rollback_version(
    card_id: str,
    version_id: str,
    *,
    confirm: bool = False,
    rebuild_indexes: bool = True,
    actor: str = "local-operator",
    actor_source: str = "local_cli",
) -> dict[str, Any]:
    """Restore a user card from an immutable version snapshot."""
    clean_card_id = str(card_id or "").strip()
    clean_version_id = str(version_id or "").strip()
    if not confirm:
        return {
            "ok": False,
            "rolled_back": False,
            "error": "confirmation_required",
            "card_id": clean_card_id,
            "version_id": clean_version_id,
        }
    row = next(
        (
            item for item in reversed(_version_rows())
            if item.get("card_id") == clean_card_id and item.get("version_id") == clean_version_id
        ),
        None,
    )
    if not row:
        raise ValueError(f"unknown knowledge version: {clean_card_id}@{clean_version_id}")
    snapshot = json.loads(Path(str(row["snapshot"])).read_text(encoding="utf-8"))
    stored_card = snapshot.get("card") or {}
    body = str(snapshot.get("body") or "")
    current = next((item for item in list_user_cards() if item.get("id") == clean_card_id), None)
    expected_hash = str((current or {}).get("body_hash") or "") if current else None
    card = import_text(
        str(stored_card.get("title") or clean_card_id),
        body,
        source_url=str(stored_card.get("source_url") or ""),
        source_type=str(stored_card.get("source_type") or "user"),
        confidence=str(stored_card.get("confidence") or ""),
        tags=_tag_list(stored_card.get("tags")),
        card_id=clean_card_id,
        license=str(stored_card.get("license") or "user_supplied"),
        expected_old_hash=expected_hash,
        actor=actor,
        actor_source=actor_source,
        version_action="rollback",
        rollback_from=clean_version_id,
    )
    indexes: dict[str, Any] = {}
    if rebuild_indexes:
        indexes["knowledge"] = rebuild_index()
        try:
            from . import retrieval
            indexes["retrieval"] = retrieval.rebuild_index()
        except Exception as exc:
            indexes["retrieval_error"] = security.redact_text(str(exc))
    return {
        "ok": True,
        "rolled_back": True,
        "card": card,
        "restored_version_id": clean_version_id,
        "created_version_id": card.get("current_version_id"),
        "indexes": indexes,
    }


def _upload_rows() -> list[dict[str, Any]]:
    p = _upload_history_file()
    if not p.exists():
        return []
    rows = []
    for line in p.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(row, dict):
            rows.append(row)
    return rows


def _upsert_upload(row: dict[str, Any]) -> None:
    with locking.exclusive_file_lock(_mutation_lock_file()):
        p = _upload_history_file()
        p.parent.mkdir(parents=True, exist_ok=True)
        rows = [r for r in _upload_rows() if r.get("id") != row.get("id")]
        rows.append(row)
        rows.sort(key=lambda r: str(r.get("created_at") or ""), reverse=True)
        _write_text_atomic(p, "\n".join(json.dumps(r, ensure_ascii=False) for r in rows) + "\n")


def list_uploads(limit: int = 50) -> dict[str, Any]:
    rows = _upload_rows()[:max(1, min(int(limit or 50), 200))]
    return {"root": str(_uploads_dir()), "uploads": rows}


def upload_detail(upload_id: str) -> dict[str, Any] | None:
    for row in _upload_rows():
        if row.get("id") == upload_id:
            return row
    return None


def list_files(limit: int = 500) -> dict[str, Any]:
    """Return user knowledge cards and uploaded source files for product UIs."""
    root = _user_base()
    uploads = []
    if _uploads_dir().exists():
        for p in sorted(_uploads_dir().rglob("*")):
            if not p.is_file():
                continue
            rel = p.relative_to(root).as_posix()
            uploads.append({
                "path": rel,
                "name": p.name,
                "size": p.stat().st_size,
                "mtime": int(p.stat().st_mtime),
                "kind": "extracted" if p.name == "extracted.md" else "raw",
            })
            if len(uploads) >= limit:
                break
    cards = []
    for card in list_user_cards():
        cards.append({
            "id": card.get("id", ""),
            "title": card.get("title", ""),
            "path": card.get("path", ""),
            "tags": list(card.get("tags") or []),
            "source_type": card.get("source_type", ""),
            "source_url": card.get("source_url", ""),
            "license": card.get("license", ""),
            "body_hash": card.get("body_hash", ""),
            "retrieved_at": card.get("retrieved_at", ""),
        })
    return {
        "root": str(root),
        "uploads_root": str(_uploads_dir()),
        "uploads": uploads,
        "cards": cards,
        "history": _upload_rows()[:max(1, min(int(limit or 500), 200))],
    }


def read_file(path: str, max_chars: int = 200_000) -> dict[str, Any]:
    target = _resolve_user_path(path)
    if not target.exists() or not target.is_file():
        raise FileNotFoundError(path)
    text = target.read_text(encoding="utf-8", errors="replace")
    truncated = len(text) > max_chars
    if truncated:
        text = text[:max_chars]
    return {
        "path": _safe_rel_path(path).as_posix(),
        "name": target.name,
        "size": target.stat().st_size,
        "mtime": int(target.stat().st_mtime),
        "content": security.redact_text(text),
        "truncated": truncated,
    }


@locking.serialized(_mutation_lock_file)
def delete_file(path: str) -> dict[str, Any]:
    rel = _safe_rel_path(path)
    target = _resolve_user_path(rel.as_posix())
    if not target.exists() or not target.is_file():
        raise FileNotFoundError(path)
    target.unlink()
    removed_card_ids = []
    if rel.parts and rel.parts[0] == "user":
        rows = []
        for card in list_user_cards():
            if card.get("path") == rel.as_posix():
                removed_card_ids.append(str(card.get("id") or ""))
            else:
                rows.append(card)
        p = _sources_file()
        p.parent.mkdir(parents=True, exist_ok=True)
        _write_text_atomic(
            p, "\n".join(json.dumps(r, ensure_ascii=False) for r in rows) + ("\n" if rows else ""),
        )
        rebuild_index()
    return {"ok": True, "path": rel.as_posix(), "removed_card_ids": removed_card_ids}


def _decode_text(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _html_to_text(text: str) -> str:
    text = re.sub(r"(?is)<(script|style|noscript|header|footer|nav)[^>]*>.*?</\1>", " ", text)
    text = re.sub(r"(?s)<br\s*/?>", "\n", text)
    text = re.sub(r"(?s)</(p|div|li|h[1-6]|tr)>", "\n", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    text = html.unescape(text)
    return re.sub(r"\n{3,}", "\n\n", re.sub(r"[ \t]+", " ", text)).strip()


def _extract_docx(data: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            xml = zf.read("word/document.xml").decode("utf-8", errors="replace")
    except Exception:
        return ""
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"<[^>]+>", "", xml)
    return html.unescape(xml).strip()


def _extract_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return ""
    lines = []
    for ws in wb.worksheets[:8]:
        lines.append(f"## {ws.title}")
        for row in ws.iter_rows(max_row=500, values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if any(vals):
                lines.append(" | ".join(vals))
    return "\n".join(lines).strip()


def _extract_pdf(data: bytes) -> str:
    reader_cls = None
    try:
        from pypdf import PdfReader
        reader_cls = PdfReader
    except Exception:
        try:
            from PyPDF2 import PdfReader
            reader_cls = PdfReader
        except Exception:
            return ""
    try:
        reader = reader_cls(io.BytesIO(data))
        pages = []
        for page in reader.pages[:80]:
            pages.append(page.extract_text() or "")
        return "\n\n".join(pages).strip()
    except Exception:
        return ""


def extract_document_text(filename: str, data: bytes) -> dict[str, Any]:
    suffix = Path(filename or "").suffix.lower()
    warnings: list[str] = []
    if suffix in {".md", ".markdown", ".txt", ".csv", ".tsv", ".json", ".yaml", ".yml", ".log"}:
        text = _decode_text(data)
    elif suffix in {".html", ".htm"}:
        text = _html_to_text(_decode_text(data))
    elif suffix == ".docx":
        text = _extract_docx(data)
        if not text:
            warnings.append("docx_text_extraction_failed")
    elif suffix in {".xlsx", ".xlsm"}:
        text = _extract_xlsx(data)
        if not text:
            warnings.append("xlsx_text_extraction_unavailable")
    elif suffix == ".pdf":
        text = _extract_pdf(data)
        if not text:
            warnings.append("pdf_text_extraction_unavailable")
    else:
        text = _decode_text(data)
        if len(text.strip()) < 20:
            warnings.append("unknown_binary_or_empty_text")
    text = security.redact_text(text).strip()
    return {"text": text, "warnings": warnings, "extension": suffix or ""}


def upload_document(filename: str, data: bytes, *, title: str = "", source_url: str = "",
                    source_type: str = "user", confidence: str = "",
                    tags: list[str] | str | None = None, card_id: str = "",
                    license: str = "user_supplied", confirm: bool = False,
                    rebuild_indexes: bool = True) -> dict[str, Any]:
    if not data:
        raise ValueError("empty upload")
    if len(data) > 25 * 1024 * 1024:
        raise ValueError("upload too large; max 25MB")
    config.ensure_dirs()
    safe_name = _safe_filename(filename)
    upload_id = f"up-{time.strftime('%Y%m%d-%H%M%S')}-{_hash(safe_name + str(len(data)) + str(time.time()))[:8]}"
    day = time.strftime("%Y%m%d")
    raw_rel = Path("uploads") / day / upload_id / safe_name
    raw_path = _user_base() / raw_rel
    raw_path.parent.mkdir(parents=True, exist_ok=True)
    raw_path.write_bytes(data)

    extracted = extract_document_text(safe_name, data)
    text = str(extracted.get("text") or "").strip()
    warnings = list(extracted.get("warnings") or [])
    if not text:
        warnings.append("no_text_extracted")
        text = f"# {title or Path(safe_name).stem}\n\n（未能自动抽取文本，请在 IvyeaOps 中补充可检索正文。）"

    clean_title = str(title or Path(safe_name).stem or "上传知识").strip()
    extracted_rel = Path("uploads") / day / upload_id / "extracted.md"
    extracted_path = _user_base() / extracted_rel
    extracted_path.parent.mkdir(parents=True, exist_ok=True)
    extracted_body = text if text.startswith("#") else f"# {clean_title}\n\n{text}"
    extracted_path.write_text(extracted_body.strip() + "\n", encoding="utf-8")

    effective_source_url = source_url or f"ivyea-upload://{upload_id}/{safe_name}"
    draft = draft_update(
        clean_title,
        extracted_body,
        source_url=effective_source_url,
        source_type=source_type or "user",
        confidence=confidence,
        tags=tags,
        card_id=card_id,
        license=license,
    )
    row = {
        "id": upload_id,
        "filename": safe_name,
        "title": clean_title,
        "raw_path": raw_rel.as_posix(),
        "extracted_path": extracted_rel.as_posix(),
        "size": len(data),
        "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "source_url": effective_source_url,
        "source_type": source_type or "user",
        "confidence": confidence or _confidence(source_type or "user"),
        "license": license or "user_supplied",
        "tags": _tag_list(tags),
        "card_id": draft.get("card_id", ""),
        "warnings": warnings,
        "text_chars": len(extracted_body),
        "body_hash": _hash(extracted_body),
    }
    _upsert_upload(row)
    result: dict[str, Any] = {
        "ok": True,
        "upload": row,
        "extraction": {
            "text_chars": len(extracted_body),
            "warnings": warnings,
            "preview": extracted_body[:1200],
        },
        "draft": draft,
    }
    if confirm:
        result["apply"] = apply_update(draft, confirm=True, rebuild_indexes=rebuild_indexes)
    return result


def apply_upload(upload_id: str, *, confirm: bool = False, rebuild_indexes: bool = True) -> dict[str, Any]:
    row = upload_detail(upload_id)
    if not row:
        raise FileNotFoundError(upload_id)
    extracted_path = _resolve_user_path(str(row.get("extracted_path") or ""))
    body = extracted_path.read_text(encoding="utf-8", errors="replace")
    draft = draft_update(
        str(row.get("title") or row.get("filename") or upload_id),
        body,
        source_url=str(row.get("source_url") or ""),
        source_type=str(row.get("source_type") or "user"),
        confidence=str(row.get("confidence") or ""),
        tags=row.get("tags") if isinstance(row.get("tags"), list) else [],
        card_id=str(row.get("card_id") or ""),
        license=str(row.get("license") or "user_supplied"),
    )
    result = apply_update(draft, confirm=confirm, rebuild_indexes=rebuild_indexes)
    if result.get("ok") and result.get("applied"):
        row["import_status"] = "imported"
        row["imported_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
        _upsert_upload(row)
    return {"ok": bool(result.get("ok")), "upload": row, "draft": draft, "result": result}


def _limited_diff(old_body: str, new_body: str, *, old_label: str, new_label: str, limit: int = 12000) -> str:
    lines = list(difflib.unified_diff(
        old_body.splitlines(),
        new_body.splitlines(),
        fromfile=old_label,
        tofile=new_label,
        lineterm="",
    ))
    text = "\n".join(lines)
    if len(text) > limit:
        return text[:limit].rstrip() + "\n... diff truncated ..."
    return text


def _safe_update_card_id(title: str, card_id: str = "") -> tuple[str, list[str]]:
    warnings = []
    raw = str(card_id or "").strip()
    if not raw:
        return f"user.{_slug(title)}", warnings
    if raw.startswith("user."):
        return raw, warnings
    warnings.append("card_id_not_user_scoped: imported as user.<id> to avoid overriding bundled knowledge")
    return f"user.{_slug(raw)}", warnings


def draft_update(title: str, body: str, *, source_url: str = "", source_type: str = "user",
                 confidence: str = "", tags: list[str] | str | None = None, card_id: str = "",
                 license: str = "user_supplied") -> dict[str, Any]:
    """Prepare a reviewed knowledge update without mutating local files."""
    clean_title = str(title or "").strip() or "用户知识"
    clean_body = security.redact_text(str(body or "")).strip()
    if not clean_body:
        raise ValueError("body is required")
    clean_body += "\n"
    clean_source_type = str(source_type or "user").strip() or "user"
    tag_list = _tag_list(tags)
    safe_id, warnings = _safe_update_card_id(clean_title, card_id)

    if not source_url:
        warnings.append("missing_source_url")
    if clean_source_type == "official" and not source_url:
        warnings.append("official_source_without_url")
    if clean_source_type.startswith("community"):
        warnings.append("community_source_requires_manual_verification")
    if len(clean_body.strip()) < 80:
        warnings.append("short_body_review_recommended")

    existing = get_card(safe_id)
    old_body = ""
    old_hash = ""
    if existing:
        old_body = security.redact_text(str(existing.get("body") or "")).strip() + "\n"
        old_hash = existing.get("body_hash") or _hash(old_body)

    new_hash = _hash(clean_body)
    if existing and old_hash == new_hash:
        action = "noop"
    elif existing:
        action = "update"
    else:
        action = "create"

    diff = "" if action == "noop" else _limited_diff(
        old_body,
        clean_body,
        old_label=f"{safe_id}:old",
        new_label=f"{safe_id}:new",
    )
    review_required = (
        clean_source_type.startswith("community")
        or not source_url
        or "short_body_review_recommended" in warnings
    )
    return {
        "ok": True,
        "action": action,
        "card_id": safe_id,
        "title": clean_title,
        "source_url": str(source_url or "").strip(),
        "source_type": clean_source_type,
        "confidence": str(confidence or _confidence(clean_source_type)),
        "license": str(license or "user_supplied"),
        "tags": tag_list,
        "old_hash": old_hash,
        "new_hash": new_hash,
        "old_scope": existing.get("scope") if existing else "",
        "diff": diff,
        "body": clean_body,
        "warnings": warnings,
        "review_required": review_required,
    }


def draft_update_from_file(path: str, *, title: str = "", source_url: str = "",
                           source_type: str = "user", confidence: str = "",
                           tags: list[str] | str | None = None, card_id: str = "",
                           license: str = "user_supplied") -> dict[str, Any]:
    p = Path(path).expanduser().resolve()
    body = p.read_text(encoding="utf-8", errors="replace")
    return draft_update(
        title or p.stem,
        body,
        source_url=source_url or str(p),
        source_type=source_type,
        confidence=confidence,
        tags=tags,
        card_id=card_id,
        license=license,
    )


def apply_update(draft: dict[str, Any], *, confirm: bool = False,
                 rebuild_indexes: bool = True) -> dict[str, Any]:
    """Apply a previously reviewed draft. Confirmation is mandatory."""
    if not isinstance(draft, dict) or not draft.get("ok"):
        return {"ok": False, "applied": False, "error": "invalid_draft"}
    if not confirm:
        return {
            "ok": False,
            "applied": False,
            "error": "confirmation_required",
            "draft": _draft_summary(draft),
        }
    if draft.get("action") == "noop":
        return {"ok": True, "applied": False, "action": "noop", "draft": _draft_summary(draft)}

    body = str(draft.get("body") or "").strip()
    if not body:
        return {"ok": False, "applied": False, "error": "draft_body_missing"}
    try:
        card = import_text(
            str(draft.get("title") or draft.get("card_id") or "用户知识"),
            body,
            source_url=str(draft.get("source_url") or ""),
            source_type=str(draft.get("source_type") or "user"),
            confidence=str(draft.get("confidence") or ""),
            tags=_tag_list(draft.get("tags")),
            card_id=str(draft.get("card_id") or ""),
            license=str(draft.get("license") or "user_supplied"),
            expected_old_hash=str(draft.get("old_hash") or ""),
            actor=str(draft.get("reviewer") or draft.get("actor") or "local-operator"),
            actor_source=str(draft.get("reviewer_source") or draft.get("actor_source") or "confirmed_update"),
        )
    except KnowledgeConflictError as exc:
        current = get_card(str(draft.get("card_id") or "")) or {}
        return {
            "ok": False,
            "applied": False,
            "error": "knowledge_update_conflict",
            "message": security.redact_text(str(exc)),
            "expected_old_hash": str(draft.get("old_hash") or ""),
            "current_hash": str(current.get("body_hash") or ""),
        }
    indexes: dict[str, Any] = {}
    if rebuild_indexes:
        indexes["knowledge"] = rebuild_index()
        try:
            from . import retrieval
            indexes["retrieval"] = retrieval.rebuild_index()
        except Exception as exc:
            indexes["retrieval_error"] = security.redact_text(str(exc))
    return {
        "ok": True,
        "applied": True,
        "action": draft.get("action"),
        "card": card,
        "indexes": indexes,
        "warnings": list(draft.get("warnings") or []),
        "version_id": card.get("current_version_id"),
        "conflicts": conflicts(),
    }


def _draft_summary(draft: dict[str, Any]) -> dict[str, Any]:
    return {
        "ok": bool(draft.get("ok")),
        "action": draft.get("action", ""),
        "card_id": draft.get("card_id", ""),
        "title": draft.get("title", ""),
        "source_type": draft.get("source_type", ""),
        "source_url": draft.get("source_url", ""),
        "old_hash": draft.get("old_hash", ""),
        "new_hash": draft.get("new_hash", ""),
        "warnings": list(draft.get("warnings") or []),
        "review_required": bool(draft.get("review_required")),
    }


def render_update_draft(draft: dict[str, Any]) -> str:
    if not draft.get("ok"):
        return f"知识更新草案生成失败：{draft.get('error', 'unknown')}"
    lines = [
        "Ivyea 知识更新草案：",
        f"- action={draft.get('action')} id={draft.get('card_id')} title={draft.get('title')}",
        f"- source={draft.get('source_type')} {draft.get('source_url') or '(missing source_url)'}",
        f"- old_hash={str(draft.get('old_hash') or '-')[:12]} new_hash={str(draft.get('new_hash') or '-')[:12]}",
        f"- review_required={bool(draft.get('review_required'))}",
    ]
    if draft.get("warnings"):
        lines.append("- warnings=" + ",".join(str(w) for w in draft.get("warnings") or []))
    diff = str(draft.get("diff") or "")
    if diff:
        lines.extend(["", diff])
    else:
        lines.append("")
        lines.append("无内容变更。")
    return "\n".join(lines)


def render_update_apply(result: dict[str, Any]) -> str:
    if not result.get("ok"):
        return f"知识更新未应用：{result.get('error', 'unknown')}"
    if not result.get("applied"):
        return f"知识更新无需应用：action={result.get('action', 'noop')}"
    card = result.get("card") or {}
    indexes = result.get("indexes") or {}
    knowledge_index = indexes.get("knowledge") or {}
    retrieval_index = indexes.get("retrieval") or {}
    return (
        f"已应用知识更新：{card.get('id')} -> {card.get('path')}\n"
        f"knowledge_index cards={knowledge_index.get('cards', '-')}\n"
        f"retrieval_index chunks={retrieval_index.get('chunks', '-')}"
    )


def import_text(title: str, body: str, *, source_url: str = "", source_type: str = "user",
                confidence: str = "", tags: list[str] | None = None, card_id: str = "",
                license: str = "user_supplied", expected_old_hash: str | None = None,
                actor: str = "local-operator", actor_source: str = "local_operation",
                version_action: str = "apply", rollback_from: str = "") -> dict[str, Any]:
    """Import a user knowledge card into ~/.ivyea/knowledge."""
    config.ensure_dirs()
    base = _user_base()
    base.mkdir(parents=True, exist_ok=True)
    safe_id = card_id or f"user.{_slug(title)}"
    rel = f"user/{_slug(safe_id)}.md"
    out = base / rel
    clean_body = security.redact_text(body).strip() + "\n"
    with locking.exclusive_file_lock(_mutation_lock_file()):
        existing = next((dict(row) for row in list_user_cards() if row.get("id") == safe_id), None)
        current_hash = str((existing or {}).get("body_hash") or "")
        if existing and not current_hash:
            try:
                current_hash = _hash(_user_base().joinpath(str(existing["path"])).read_text(encoding="utf-8"))
            except OSError:
                current_hash = ""
        if expected_old_hash is not None and current_hash != str(expected_old_hash):
            raise KnowledgeConflictError(
                f"knowledge update conflict for {safe_id}: expected {expected_old_hash or '-'}, current {current_hash or '-'}"
            )

        parent_version_id = str((existing or {}).get("current_version_id") or "")
        revision = int((existing or {}).get("revision") or 0)
        if existing and not parent_version_id:
            baseline_body = _user_base().joinpath(str(existing["path"])).read_text(
                encoding="utf-8", errors="replace",
            )
            revision = max(1, revision)
            parent_version_id = _new_version_id(safe_id, revision, _hash(baseline_body))
            baseline = {
                **existing,
                "revision": revision,
                "current_version_id": parent_version_id,
                "parent_version_id": "",
                "body_hash": _hash(baseline_body),
            }
            _record_version_unlocked(
                baseline,
                baseline_body,
                action="baseline",
                actor="system-migration",
                actor_source="version_bootstrap",
            )

        revision = revision + 1 if existing else 1
        body_hash = _hash(clean_body)
        version_id = _new_version_id(safe_id, revision, body_hash)
        card = {
            "id": safe_id,
            "title": title.strip() or safe_id,
            "category": "user",
            "source_type": source_type or "user",
            "confidence": confidence or _confidence(source_type or "user"),
            "retrieved_at": time.strftime("%Y-%m-%d"),
            "license": license or "user_supplied",
            "source_url": source_url,
            "path": rel,
            "tags": tags or [],
            "scope": "user",
            "body_hash": body_hash,
            "revision": revision,
            "current_version_id": version_id,
            "parent_version_id": parent_version_id,
            "updated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        }
        _write_text_atomic(out, clean_body)
        _upsert_source_unlocked(card)
        _record_version_unlocked(
            card,
            clean_body,
            action=version_action,
            actor=actor,
            actor_source=actor_source,
            rollback_from=rollback_from,
        )
        return card


def annotate_user_card(card_id: str, metadata: dict[str, Any]) -> dict[str, Any]:
    """Attach allowlisted structured metadata to an existing user knowledge card."""
    card = next((dict(row) for row in list_user_cards() if row.get("id") == card_id), None)
    if not card:
        raise FileNotFoundError(card_id)
    allowed = {
        "marketplaces", "locales", "evidence_id", "evidence_kind", "observed_at",
        "captured_at", "diagnostic", "authority_tier", "evidence_class", "source_quality",
    }
    for key, value in metadata.items():
        if key in allowed:
            card[key] = value
    _upsert_source(card)
    return card


def _import_destination(namespace: str, rel: Path) -> tuple[str, str]:
    safe_namespace = _safe_path_segment(namespace or "imported", "imported")
    parts = [_safe_path_segment(part, "dir") for part in rel.with_suffix("").parts]
    if not parts:
        parts = ["item"]
    file_stem = parts[-1]
    dir_parts = parts[:-1]
    body_key = rel.as_posix()
    card_id = f"user.{safe_namespace}.{_hash(body_key)[:16]}"
    out_rel = Path("user") / "imported" / safe_namespace / Path(*dir_parts) / f"{file_stem}.md"
    return card_id, out_rel.as_posix()


def _scan_import_file(root: Path, path: Path, *, namespace: str, max_file_bytes: int) -> dict[str, Any]:
    rel = path.relative_to(root)
    rel_text = rel.as_posix()
    size = path.stat().st_size
    base = {
        "source_path": rel_text,
        "size": size,
        "extension": path.suffix.lower(),
    }
    if any(part.startswith(".") or part in IGNORED_IMPORT_DIRS for part in rel.parts):
        return {**base, "importable": False, "reason": "hidden_or_ignored_path"}
    if path.suffix.lower() not in IMPORTABLE_SUFFIXES:
        return {**base, "importable": False, "reason": "unsupported_extension"}
    if size > max_file_bytes:
        return {**base, "importable": False, "reason": "file_too_large"}
    data = path.read_bytes()
    extracted = extract_document_text(path.name, data)
    body = str(extracted.get("text") or "").strip()
    if not body:
        return {**base, "importable": False, "reason": "no_text_extracted", "warnings": extracted.get("warnings") or []}
    if not body.startswith("#"):
        body = f"# {_first_heading_or_stem(body, path.stem)}\n\n{body}"
    body = body.strip() + "\n"
    card_id, target_path = _import_destination(namespace, rel)
    existing = get_card(card_id)
    body_hash = _hash(body)
    old_hash = str(existing.get("body_hash") or "") if existing else ""
    if existing and old_hash == body_hash:
        action = "noop"
    elif existing:
        action = "update"
    else:
        action = "create"
    return {
        **base,
        "importable": True,
        "action": action,
        "card_id": card_id,
        "title": _first_heading_or_stem(body, path.stem),
        "target_path": target_path,
        "source_url": f"{namespace}://{rel_text}",
        "tags": [namespace, *[str(p) for p in rel.parts[:-1]]],
        "warnings": extracted.get("warnings") or [],
        "text_chars": len(body),
        "body_hash": body_hash,
        "old_hash": old_hash,
        "body": body,
    }


def import_directory(root: str, *, namespace: str = "gbrain", confirm: bool = False,
                     max_files: int = 1000, max_file_bytes: int = DEFAULT_IMPORT_FILE_BYTES,
                     rebuild_indexes: bool = True) -> dict[str, Any]:
    """Scan or import a legacy markdown knowledge directory into user knowledge.

    This is intentionally file-based and does not depend on the old GBrain CLI.
    It lets IvyeaOps inherit ~/brain-style knowledge into ~/.ivyea/knowledge.
    """
    root_path = Path(root or "").expanduser().resolve()
    if not root_path.exists() or not root_path.is_dir():
        raise FileNotFoundError(str(root_path))
    safe_namespace = _safe_path_segment(namespace or "gbrain", "gbrain")
    limit = max(1, min(int(max_files or 1000), 5000))
    byte_limit = max(1024, min(int(max_file_bytes or DEFAULT_IMPORT_FILE_BYTES), 25 * 1024 * 1024))
    scanned = 0
    candidates: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for path in sorted(root_path.rglob("*")):
        if not path.is_file():
            continue
        try:
            rel = path.relative_to(root_path)
        except ValueError:
            continue
        if any(part.startswith(".") or part in IGNORED_IMPORT_DIRS for part in rel.parts):
            continue
        scanned += 1
        row = _scan_import_file(root_path, path, namespace=safe_namespace, max_file_bytes=byte_limit)
        if row.get("importable"):
            public_row = {k: v for k, v in row.items() if k != "body"}
            candidates.append(public_row)
        else:
            skipped.append(row)
        if len(candidates) >= limit:
            break

    imported: list[dict[str, Any]] = []
    unchanged: list[dict[str, Any]] = []
    if confirm:
        base = _user_base()
        for row in candidates:
            if row.get("action") == "noop":
                unchanged.append(row)
                continue
            source_file = root_path / str(row["source_path"])
            body = _scan_import_file(root_path, source_file, namespace=safe_namespace, max_file_bytes=byte_limit).get("body", "")
            if not body:
                continue
            target = (base / str(row["target_path"])).resolve()
            if base.resolve() not in target.parents:
                raise ValueError("import target escapes knowledge directory")
            target.parent.mkdir(parents=True, exist_ok=True)
            clean_body = security.redact_text(str(body)).strip() + "\n"
            target.write_text(clean_body, encoding="utf-8")
            card = {
                "id": row["card_id"],
                "title": row.get("title") or row["card_id"],
                "category": f"legacy_{safe_namespace}",
                "source_type": f"legacy_{safe_namespace}",
                "confidence": "user_supplied",
                "retrieved_at": time.strftime("%Y-%m-%d"),
                "license": "user_supplied",
                "source_url": row.get("source_url") or "",
                "path": row["target_path"],
                "tags": row.get("tags") or [safe_namespace],
                "scope": "user",
                "body_hash": _hash(clean_body),
            }
            _upsert_source(card)
            imported.append({**row, "card": card})

    indexes: dict[str, Any] = {}
    if confirm and rebuild_indexes:
        indexes["knowledge"] = rebuild_index()
    return {
        "ok": True,
        "root": str(root_path),
        "namespace": safe_namespace,
        "confirm": bool(confirm),
        "scanned_files": scanned,
        "candidates": candidates,
        "skipped": skipped[:200],
        "summary": {
            "candidate_files": len(candidates),
            "skipped_files": len(skipped),
            "create": len([r for r in candidates if r.get("action") == "create"]),
            "update": len([r for r in candidates if r.get("action") == "update"]),
            "noop": len([r for r in candidates if r.get("action") == "noop"]),
            "imported": len(imported),
            "unchanged": len(unchanged),
            "limit_reached": len(candidates) >= limit,
        },
        "imported": imported,
        "unchanged": unchanged,
        "indexes": indexes,
    }


def import_file(path: str, *, title: str = "", source_type: str = "user",
                confidence: str = "", tags: list[str] | None = None, card_id: str = "",
                license: str = "user_supplied") -> dict[str, Any]:
    p = Path(path).expanduser().resolve()
    body = p.read_text(encoding="utf-8", errors="replace")
    return import_text(
        title or p.stem,
        body,
        source_url=str(p),
        source_type=source_type,
        confidence=confidence,
        tags=tags,
        card_id=card_id,
        license=license,
    )


def import_url(url: str, *, title: str = "", source_type: str = "user",
               confidence: str = "", tags: list[str] | None = None, card_id: str = "",
               license: str = "user_supplied") -> dict[str, Any]:
    import httpx

    r = httpx.get(url, timeout=30, follow_redirects=True, headers={"User-Agent": "ivyea-agent/0.4"})
    r.raise_for_status()
    text = r.text
    if "html" in r.headers.get("content-type", ""):
        text = re.sub(r"(?is)<(script|style)[^>]*>.*?</\1>", "", text)
        text = re.sub(r"(?s)<[^>]+>", " ", text)
        text = re.sub(r"[ \t]+", " ", re.sub(r"\n\s*\n+", "\n", text)).strip()
    return import_text(
        title or url,
        text,
        source_url=url,
        source_type=source_type,
        confidence=confidence,
        tags=tags,
        card_id=card_id,
        license=license,
    )


def _upsert_source_unlocked(card: dict[str, Any]) -> None:
    p = _sources_file()
    p.parent.mkdir(parents=True, exist_ok=True)
    rows = [c for c in list_user_cards() if c.get("id") != card["id"]]
    rows.append(card)
    _write_text_atomic(p, "\n".join(json.dumps(r, ensure_ascii=False) for r in rows) + "\n")


def _upsert_source(card: dict[str, Any]) -> None:
    with locking.exclusive_file_lock(_mutation_lock_file()):
        _upsert_source_unlocked(card)


@locking.serialized(_mutation_lock_file)
def rebuild() -> dict[str, Any]:
    """Validate user knowledge metadata and prune rows with missing files."""
    rows = []
    missing = []
    for card in list_user_cards():
        if _user_base().joinpath(card["path"]).exists():
            rows.append(card)
        else:
            missing.append(card["id"])
    p = _sources_file()
    p.parent.mkdir(parents=True, exist_ok=True)
    _write_text_atomic(
        p, "\n".join(json.dumps(r, ensure_ascii=False) for r in rows) + ("\n" if rows else ""),
    )
    idx = rebuild_index()
    return {"user_cards": len(rows), "missing_pruned": missing, "sources": str(p), "index": idx}


def _conn() -> sqlite3.Connection:
    p = _index_file()
    p.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(p))
    conn.row_factory = sqlite3.Row
    conn.execute("""CREATE TABLE IF NOT EXISTS cards (
        id TEXT PRIMARY KEY,
        title TEXT,
        category TEXT,
        scope TEXT,
        source_type TEXT,
        confidence TEXT,
        freshness TEXT,
        source_quality TEXT,
        retrieved_at TEXT,
        license TEXT,
        body_hash TEXT,
        tags TEXT,
        source_url TEXT,
        body TEXT
    )""")
    _ensure_column(conn, "cards", "category", "TEXT")
    _ensure_column(conn, "cards", "freshness", "TEXT")
    _ensure_column(conn, "cards", "source_quality", "TEXT")
    return conn


def _ensure_column(conn: sqlite3.Connection, table: str, column: str, decl: str) -> None:
    cols = {row[1] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {decl}")


def _fts_ok(conn: sqlite3.Connection) -> bool:
    try:
        conn.execute("CREATE VIRTUAL TABLE IF NOT EXISTS cards_fts USING fts5(id, title, tags, body)")
        return True
    except Exception:
        return False


def rebuild_index() -> dict[str, Any]:
    conn = _conn()
    fts = _fts_ok(conn)
    conn.execute("DELETE FROM cards")
    if fts:
        conn.execute("DELETE FROM cards_fts")
    count = 0
    for card in list_cards():
        body = _read_body(card)
        body_hash = card.get("body_hash") or _hash(body)
        tags = ",".join(card.get("tags") or [])
        conn.execute(
            "INSERT OR REPLACE INTO cards "
            "(id,title,category,scope,source_type,confidence,freshness,source_quality,retrieved_at,license,body_hash,tags,source_url,body) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                card["id"], card.get("title", ""), card.get("category", ""),
                card.get("scope", "builtin"),
                card.get("source_type", ""), card.get("confidence", ""),
                card.get("freshness", ""), card.get("source_quality", ""),
                card.get("retrieved_at", ""), card.get("license", ""),
                body_hash, tags, card.get("source_url", ""), body,
            ),
        )
        if fts:
            conn.execute("INSERT INTO cards_fts (id, title, tags, body) VALUES (?,?,?,?)",
                         (card["id"], card.get("title", ""), tags, body))
        count += 1
    conn.commit()
    conn.close()
    return {"cards": count, "fts": fts, "db": str(_index_file()), "source_registry": source_registry()["summary"]}


def search_index(query: str, limit: int = 5) -> list[dict[str, Any]]:
    if not _index_file().exists():
        rebuild_index()
    conn = _conn()
    rows = []
    try:
        if _fts_ok(conn):
            rows = conn.execute(
                "SELECT c.* FROM cards_fts f JOIN cards c ON c.id=f.id "
                "WHERE cards_fts MATCH ? LIMIT ?",
                (query, limit),
            ).fetchall()
        if not rows:
            like = f"%{query}%"
            rows = conn.execute(
                "SELECT * FROM cards WHERE id LIKE ? OR title LIKE ? OR tags LIKE ? OR body LIKE ? LIMIT ?",
                (like, like, like, like, limit),
            ).fetchall()
    except Exception:
        like = f"%{query}%"
        rows = conn.execute(
            "SELECT * FROM cards WHERE id LIKE ? OR title LIKE ? OR tags LIKE ? OR body LIKE ? LIMIT ?",
            (like, like, like, like, limit),
        ).fetchall()
    conn.close()
    out = []
    for row in rows:
        d = dict(row)
        d["snippet"] = _snippet(d.get("body", ""), re.findall(r"[\w\u4e00-\u9fff+.-]+", query))
        d["tags"] = [t for t in (d.get("tags") or "").split(",") if t]
        out.append(d)
    return out


def conflicts() -> list[dict[str, Any]]:
    cards = list_cards()
    official = [c for c in cards if c.get("source_type") == "official" or str(c.get("source_type", "")).startswith("official_plus")]
    user = [c for c in cards if c.get("scope") == "user"]
    rows = []
    seen: set[str] = set()
    generic_overlap_tags = {
        "amazon", "compliance", "gbrain", "marketplace", "official",
        "operations", "policy", "seller", "seller-central",
    }

    def add(card: dict[str, Any], level: str, reason_code: str, reason: str, related: list[str] | None = None) -> None:
        fingerprint = _hash("|".join([str(card.get("id")), reason_code, *(related or [])]))[:16]
        if fingerprint in seen:
            return
        seen.add(fingerprint)
        row = {
            "fingerprint": fingerprint,
            "level": level,
            "id": card["id"],
            "reason_code": reason_code,
            "reason": reason,
            "related": (related or [])[:5],
            "review_required": True,
        }
        rows.append(row)

    for card in user:
        if not card.get("license"):
            add(card, "warn", "missing_license", "用户知识卡缺 license")
        tags = {str(tag).lower() for tag in card.get("tags") or []}
        conflict_tags = tags - generic_overlap_tags
        body = _read_body(card).lower()
        overlaps = [
            row["id"] for row in official
            if conflict_tags and conflict_tags.intersection({str(tag).lower() for tag in row.get("tags") or []})
        ]
        reverse = any(k in body for k in ("不要", "禁止", "不建议", "avoid", "do not", "never"))
        if reverse and overlaps:
            add(
                card, "review", "directional_claim_overlap",
                "用户/社区知识含反向表述，且标签与官方知识重叠；需要人工确认是否冲突", overlaps,
            )
        undocumented_algorithm = any(term in body for term in (
            "流量池", "算法权重", "隐藏权重", "固定权重", "traffic pool", "hidden weight", "ranking weight",
        ))
        universal_claim = any(term in body for term in (
            "一定", "必然", "保证", "永远", "固定为", "guarantee", "always", "must", "will always",
        ))
        numeric_rule = bool(re.search(r"\b\d+(?:\.\d+)?\s*(?:%|clicks?|days?)\b|\d+(?:\.\d+)?\s*(?:次点击|天)", body))
        if overlaps and (undocumented_algorithm or (universal_claim and numeric_rule)):
            add(
                card, "review", "unsupported_algorithm_or_numeric_claim",
                "用户/旧知识包含未公开算法或绝对数值规则，且与官方主题重叠；必须降级为假设或补充当前证据", overlaps,
            )
        source_url = str(card.get("source_url") or "")
        if str(card.get("source_type") or "") == "official" and not source_url.startswith((
            "https://advertising.amazon.com/", "https://sell.amazon.",
            "https://sellercentral.amazon.", "https://developer-docs.amazon",
        )):
            add(card, "fail", "official_provenance_invalid", "标记为 official 的用户知识没有可验证的 Amazon 官方 URL")
    rows.sort(key=lambda row: ({"fail": 0, "review": 1, "warn": 2}.get(row["level"], 3), row["id"], row["reason_code"]))
    return rows


def render_conflicts() -> str:
    rows = conflicts()
    if not rows:
        return "Ivyea 知识库冲突审计\n\nOK 未发现明显冲突风险。\n"
    lines = ["Ivyea 知识库冲突审计", ""]
    for r in rows:
        related = f" related={','.join(r.get('related', []))}" if r.get("related") else ""
        lines.append(f"- [{r['level']}] {r['id']}: {r['reason']}{related}")
    return "\n".join(lines) + "\n"
